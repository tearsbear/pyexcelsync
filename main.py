import openpyxl
from pandas_datareader import data, wb
import xlrd
import time


def add_column(column):
    wb = openpyxl.load_workbook("works.xlsx")
    ws = wb["Sheet1"]

    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=2, value=value)
    wb.save("output.xlsx")
    print("Success...")
    print("---- Check output.xlsx file for the result!")


def main():
    # change your file work here, it must be .xlsx extension!
    file_work = "works.xlsx"

    # header name for no surat
    dataSurat = []
    dataSurat.append("NoSurat")

    print("---- reading and opening file")
    time.sleep(3)
    wb = openpyxl.load_workbook(file_work)
    ws = wb["Sheet1"]
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            filename = cell.value
            removeEx = cell.value.replace(".xls", "")
            foldername = removeEx.replace("LAMPIRAN_", "")
            workbook = xlrd.open_workbook_xls(
                f"file/DOKUMEN_{foldername}/{filename}", ignore_workbook_corruption=True
            )
            sheet = workbook.sheet_by_index(0)
            data_0 = []
            for i in range(sheet.nrows):
                data_0.append(sheet.cell(i, 0).value)
            no_surat = data_0[0].replace("NO SURAT SIK: ", "")
            print(f"NO SIK for {filename}: \n{no_surat}")
            dataSurat.append(no_surat)
    print(f"---- create new file based on {file_work}")
    time.sleep(3)
    print("---- adding result as new column")
    time.sleep(3)
    add_column(dataSurat)


if __name__ == "__main__":
    main()
